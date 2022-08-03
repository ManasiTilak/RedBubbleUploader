from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import glob
import os

#setting webdriver and removing *Chrome is being controlled by automated test software* infobar
options = webdriver.ChromeOptions()
options.add_experimental_option("useAutomationExtension", False)
options.add_experimental_option("excludeSwitches",["enable-automation"])

driver_path = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(executable_path=driver_path, chrome_options=options)
#set driver
driver.get("https://www.redbubble.com/auth/login")
#OpenPyxl ki workbook
wb = openpyxl.load_workbook(r"C:\Users\salil\Desktop\RedbubbleData1.xlsx")
#Workbook ki sheet ka obj
obj1 = wb['RedbubbleData']
#Get max rows in excelsheet
row = obj1.max_row
#Setting main folder and image pattern
src_folder = r"C:\Users\salil\Desktop\RedbubbleProgTry"
pattern = "\*.png"
#Set dir to an array of sort containing files with pattern : .png
list_File = glob.glob(src_folder+pattern)
def PerformSignIn():
    #getting username
    username = driver.find_element(By.ID, "ReduxFormInput1")
    username.send_keys("manasitilak12@gmail.com")

    #getting password
    passwd = driver.find_element(By.ID, "ReduxFormInput2")
    passwd.send_keys("Sana2020")
    passwd.send_keys(Keys.RETURN)
    time.sleep(10)


def Navi(f, j):
    driver.get("https://www.redbubble.com/portfolio/images/new")
    driver.find_element(By.CLASS_NAME, "copy-icon").click()

    goto_Button = driver.find_element(By.CLASS_NAME,"works_work-menu-link")
    goto_Button.click()
    time.sleep(10)
    #navigating the drop-down to Copy settings.
    try:
        optionsMenus = driver.find_elements(By.CLASS_NAME, "works_work-menu")
        for eachMenu in optionsMenus:
                anchors = eachMenu.find_elements(By.TAG_NAME, "a")
                anchors[2].click()
        time.sleep(10)

    except Exception as e:
        print(e)

    time.sleep(10)

    # clearing input fields
    title = driver.find_element(By.ID, "work_title_en")
    title.clear()

    tit = obj1.cell(j, 1).value
    title.send_keys(tit)
    tagss = driver.find_element(By.ID, "work_tag_field_en")
    tagss.clear()
    tg = obj1.cell(j, 2).value
    tagss.send_keys(tg)
    desc = driver.find_element(By.ID, "work_description_en")
    desc.clear()
    des = obj1.cell(j, 3).value
    desc.send_keys(des)
    # os.chdir(r'C:\Users\salil\Desktop\Redbubble\NewShop_PetBandanas\Uploaded_Petbandana')
    desc = driver.find_element(By.ID, "select-image-base")
    desc.send_keys(f)

    time.sleep(50)
    driver.find_element(By.ID, "rightsDeclaration").click()
    time.sleep(15)
    driver.find_element(By.ID, "submit-work").click()
    #setting wait until to start new upload
    urlpattern = 'https://www.redbubble.com/studio/promote'
    wait = WebDriverWait(driver, 50)
    wait.until(EC.url_contains(urlpattern))

    print("Yayyy....Done. Now What??")
    # driver.get("https://www.redbubble.com/portfolio/images/new")

# class waittest:
#     def __init__(self, locator, attr, value):
#         self._locator = locator
#         self._attribute = attr
#         self._attribute_value = value
#
#     def __call__(self, driver):
#         element = driver.find_element_by_xpath(self._locator)
#         if element.get_attribute(self._attribute) == self._attribute_value:
#             return element
#         else:
#             return False

def closepop_up():
    popup_Button =  driver.find_element(By.XPATH,'/html/body/div[1]/div[7]/button')
    popup_Button.click()

def getnumofimages():
    len = 0
    # Iterate directory
    for path in os.listdir(src_folder):
        # check if current path is a file
        if os.path.isfile(os.path.join(src_folder, path)):
            len += 1
    getdata(len)

def getdata(count):
    for i in range(2, row + 1):
        for j in range(0, count):
            if (i - 2) is j:
                Navi(list_File[j],i)

if __name__ == '__main__':
    PerformSignIn()
    getnumofimages()

    # files = glob.glob(src_folder + pattern)
    # row = obj1.max_row
    # for file in files:
    #     for i in range(2, row + 1):
    #         if file is i:
    #             Navi(file, i)

    time.sleep(10)
    driver.quit()