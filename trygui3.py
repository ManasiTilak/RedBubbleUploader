from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QLabel, QFileDialog, QDialog, QLineEdit
from PyQt5 import uic
import sys
import os
from openpyxl import load_workbook

wb = load_workbook("getuserdatatry.xlsx")
obj = wb['datax']

userdata = load_workbook("StoreuserData.xlsx")
userobj = userdata['Sheet1']

class UserInfoUI(QMainWindow):
    def __init__(self):
        super(UserInfoUI, self).__init__()
        #loading the GUI Window
        uic.loadUi("usernamegui.ui", self)
        #find user name and password labels 
        self.usernm = self.findChild(QLineEdit, "usernameinp")
        self.passwrd = self.findChild(QLineEdit, "passwordinp")

# UpdateUser
        # Find the two buttons
        self.GotoNext = self.findChild(QPushButton, "pushButton")
        self.UpdateUser = self.findChild(QPushButton, "updateinfo")
        #initialize the next window,ui
        self.nextwin = UI()
        
        # define what happens when buttons are clicked
        self.UpdateUser.clicked.connect(self.saveinfonav)
        self.GotoNext.clicked.connect(self.nextwin.show)

    def saveinfonav(self):
        #Getting username
        userusername = self.usernm.text()
        print(str(userusername))

        #Getting username
        userpassword = self.passwrd.text()
        print(str(userpassword))

        #saving user info to a excelfile
        userobj['A1'].value = str(userusername)
        userobj['B1'].value = str(userpassword)
        
        userdata.save("StoreuserData.xlsx")

        


class DialogWin(QDialog):
    def __init__(self):
        super(DialogWin, self).__init__()
        #loading the DialogGUI Window
        uic.loadUi("endscreen.ui", self)

class UI(QMainWindow):

      
    def __init__(self):
        super(UI, self).__init__()

        #loading the GUI Window
        uic.loadUi("betterdialog.ui", self)

        #for the image
        self.browseImage = self.findChild(QPushButton, "BrowseImg")      
        self.labelIm = self.findChild(QLabel, "imagFolder")
        self.browseImage.clicked.connect(self.imgclicker)

        #for the excel file
        self.browseEx = self.findChild(QPushButton, "browseIxl")      
        self.labelXl = self.findChild(QLabel, "excelFile")
        self.browseEx.clicked.connect(self.xlclicker)

        #for done btn to open second window
        self.doneButton = self.findChild(QPushButton, "doneBtn")
        self.newDialog = DialogWin()

        # if self.labelIm.text is "Uploaded Image Folder" and self.labelXl.text is "Uploaded Excel File":
            
        #     print("label image is open")
        self.doneButton.clicked.connect(self.newDialog.show)
        # else:
        #     self.doneButton.clicked.connect(self.datawarning)
                
    # def datawarning(self):
        # print("warning")

    def imgclicker(self):
        #Open file dialog. Returns a tuple
        fname = QFileDialog.getOpenFileName(self, "Open File", "", "PNG Files (*.png);;JPG Files(*.jpg)")

        #output filename to screen and save to application data:

        if fname:
            self.labelIm.setText("Uploaded Image Folder")
            a2 = obj['A2']
            a2.value = str(fname[0])
            wb.save("getuserdatatry.xlsx")


    def xlclicker(self):
        xname = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx))")

        #output xcel filename to the screen and save to application data:
        if xname:
            self.labelXl.setText("Uploaded Excel File")
            b2 = obj['B2']
            b2.value = xname[0]
            wb.save("getuserdatatry.xlsx")


def main():
    app = QApplication(sys.argv)
    main = UserInfoUI()
    main.show()
    sys.exit(app.exec_())
        
            
if __name__ == "__main__":
    main()

    