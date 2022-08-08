from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import glob
import os
# import trygui3
# from trygui3 import UI

#setting webdriver and removing *Chrome is being controlled by automated test software* infobar
options = webdriver.ChromeOptions()
options.add_experimental_option("useAutomationExtension", False)
options.add_experimental_option("excludeSwitches",["enable-automation"])

driver_path = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(executable_path=driver_path, chrome_options=options)
#set driver
driver.get("https://www.redbubble.com/auth/login")



def getdir():
    #this function reads the image address from getdir.txt and passes it to strmani
    with open("getdir.txt", "r") as f1:
        string_mani(f1.read())


def string_mani(str):
    originalString = str
    replacedString = originalString.replace('/', '\\')
    reversedString = replacedString[::-1]
    slashpos = reversedString.find('\\')
    appendString = reversedString[slashpos+1:]
    finalString = appendString [::-1]
    print(f'final string>>>>>>is >>>>>{finalString}')
    # #Setting main folder and image pattern
    src_folder = finalString
    getnumofimages(src_folder)


def performSignIn():
    #sending username
    username = driver.find_element(By.ID, "ReduxFormInput1")
    username.send_keys("manasitilak12@gmail.com")

    #sending password
    passwd = driver.find_element(By.ID, "ReduxFormInput2")
    passwd.send_keys("Sana2020")
    passwd.send_keys(Keys.RETURN)
    time.sleep(10)


def navi(f, j,obj):
#function being called, it's filling in image and row specific data in the given form
    driver.get("https://www.redbubble.com/portfolio/images/new")
    driver.find_element(By.CLASS_NAME, "copy-icon").click()

    goto_Button = driver.find_element(By.CLASS_NAME,"works_work-menu-link")
    goto_Button.click()
    time.sleep(10)
#navigating the drop-down to Copy settings.
    try:
        optionsMenus = driver.find_elements(By.CLASS_NAME, "works_work-menu")
        for eachMenu in optionsMenus:
                anchors = eachMenu.find_elements(By.TAG_NAME, "a")
                anchors[2].click()
        time.sleep(10)

    except Exception as e:
        print(e)

    time.sleep(10)

    # Title
    title = driver.find_element(By.ID, "work_title_en")
    title.clear()
    tit = obj.cell(j, 1).value
    title.send_keys(tit)
    #Tags
    tagss = driver.find_element(By.ID, "work_tag_field_en")
    tagss.clear()
    tg = obj.cell(j, 2).value
    tagss.send_keys(tg)
    #description
    desc = driver.find_element(By.ID, "work_description_en")
    desc.clear()
    des = obj.cell(j, 3).value
    desc.send_keys(des)
    #image
    img = driver.find_element(By.ID, "select-image-base")
    img.send_keys(f)
    #waiting for the image to be uploaded. I need a wait until here
    # time.sleep(50)
    # single-upload with-uploader
    #clicking remaining 2 boxes
    driver.find_element(By.ID, "rightsDeclaration").click()
    time.sleep(50)
    driver.find_element(By.ID, "submit-work").click()
    #setting wait until to start new upload
    urlpattern = 'https://www.redbubble.com/studio/promote'
    wait = WebDriverWait(driver, 50)
    wait.until(EC.url_contains(urlpattern))

def getnumofimages(folder):
    #Getting total number of images in the dir
    src_folder = folder
    len = 0

    # Set dir to an array of sort containing files with pattern : .png
    pattern = "\*.png"
    # Setting Glob
    list_File = glob.glob(src_folder+pattern)

    # Iterate directory
    for path in os.listdir(src_folder):
        # check if current path is a file
        if os.path.isfile(os.path.join(src_folder, path)):
            len += 1
    print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    print(list_File)
    getdata(len, list_File)

def getdata(count, listF):
    #this sends data to func navi
    #it sends one image and the row corresponding to that image.
    #it keeps sending till all the images are done
    #OpenPyxl ki workbook
    wb = openpyxl.load_workbook(r"C:\Users\salil\Desktop\RedbubbleData1.xlsx")
    #Workbook ki sheet ka obj
    obj1 = wb['RedbubbleData']
    #Get max rows in excelsheet
    row = obj1.max_row
    
    list_Fi = listF
    for row in range(2, row + 1):
        for picnum in range(0, count):
            if (row - 2) is picnum:
                navi(list_Fi[picnum],row,obj1)

if __name__ == '__main__':
    performSignIn()
    getdir()
    time.sleep(10)
    driver.quit()
    getdir()